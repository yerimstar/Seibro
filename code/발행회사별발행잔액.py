import requests
import json
import xmltodict
from openpyxl import load_workbook

date = input("날짜를 입력하세요 ex) 20210320 : ")
url = "https://seibro.or.kr/websquare/engine/proworks/callServletService.jsp"

payload = "<reqParam action=\"issucoByIssuRemaListEL1\" task=\"ksd.safe.bip.cnts.DerivCombi.process.DeriCommPTask\"><STD_DT value=\"" +date+ "\"/></reqParam>"
headers = {
  'Accept': 'application/xml',
  'Accept-Encoding': 'gzip, deflate, br',
  'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7,zh-CN;q=0.6,zh;q=0.5,zh-TW;q=0.4,zh-HK;q=0.3',
  'Connection': 'keep-alive',
  'Content-Length': '138',
  'Content-Type': 'application/xml; charset="UTF-8"',
  'Cookie': 'WMONID=DZISaX0l1rM; JSESSIONID=YHpGONfyKAr9KK97mJUWtXFQByrC5H9HnVHXMVjOBOFxp39G_mhZ!293748093; lastAccess=1620380805857; JSESSIONID=m8BULWAFiVAIxwy-vcynVWzcuhNCXRPQk18DF4OycGxnfTA1ZRyD!-1270617111; WMONID=jzwO4AEb8k2',
  'Host': 'seibro.or.kr',
  'Origin': 'https://seibro.or.kr',
  'Referer': 'https://seibro.or.kr/websquare/control.jsp?w2xPath=/IPORTAL/user/derivCombi/BIP_CNTS07003V.xml&menuNo=193',
  'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
  'sec-ch-ua-mobile': '?0',
  'Sec-Fetch-Dest': 'empty',
  'Sec-Fetch-Mode': 'cors',
  'Sec-Fetch-Site': 'same-origin',
  'submissionid': 'submission_issucoByIssuRemaListEL1',
  'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.85 Safari/537.36'
}

response = requests.request("POST", url, headers=headers, data=payload)

xmlString = response.text
jsonDump = json.dumps(xmltodict.parse(xmlString),ensure_ascii=False)
jsonText = json.loads(jsonDump)

wb = load_workbook(filename = '/Users/yerimstar/PycharmProjects/예탁원/템플릿/예탁원_발행회사별발행잔액_템플릿.xlsm', data_only=True)
ws = wb[wb.sheetnames[0]]
length = len(jsonText["vector"]["data"])

if length == 23:
  for i in range(0, 23):
    print(jsonText["vector"]["data"][i]["result"]["REP_SECN_NM"]["@value"])
    value = str(jsonText["vector"]["data"][i]["result"]["REP_SECN_NM"]["@value"])
    ws.cell(row = i + 4,column = 1).value = value
    ws.cell(row = i + 4, column = 2).value = "발행잔액"
    for j in range(1,25):
      cellNum = "ROW2_COL"+str(j)
      cellValue = jsonText["vector"]["data"][i]["result"][cellNum]["@value"]
      ws.cell(row = i + 4, column = j + 2).value = cellValue
    totalValue = jsonText["vector"]["data"][i]["result"]["ROW2_HAP"]["@value"]
    ws.cell(row = i + 4, column = 27).value = totalValue
  wb.save("/Users/yerimstar/PycharmProjects/예탁원/결과물/예탁원_발행회사별발행잔액_("+date+").xlsx")
else:
  print("값이 없습니다.")



