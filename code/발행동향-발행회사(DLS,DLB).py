import requests
import json
import xmltodict
from openpyxl import load_workbook

value1 = input("상품유형을 입력하세요 ex) 전체, DLS, DLB중 1개 입력 : ")
date1 = input("시작 날짜를 입력하세요 ex) 20210501 : ")
date2 = input("종료 날짜를 입력하세요 ex) 20210510 : ")

url = "https://seibro.or.kr/websquare/engine/proworks/callServletService.jsp"

if value1 == "전체":
    value1 = "전체(DLS,DLB)_"
    num1 = "99"
elif value1 == "DLS":
    num1 = "43"
elif value1 == "DLB":
    num1 = "43-B"


payload = "<reqParam action=\"issucoTop10List\" task=\"ksd.safe.bip.cnts.DerivCombi.process.DeriDLSPTask\"><SECN_TPCD value=\""+num1+"\"/><DERI_INFO value=\"2\"/><STD_DT1 value=\""+date1+"\"/><STD_DT2 value=\""+date2+"\"/><START_PAGE value=\"1\"/><END_PAGE value=\"10\"/></reqParam>"
headers = {
  'Accept': 'application/xml',
  'Accept-Encoding': 'gzip, deflate, br',
  'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7,zh;q=0.6,zh-CN;q=0.5,zh-TW;q=0.4,zh-HK;q=0.3',
  'Connection': 'keep-alive',
  'Content-Type': 'application/xml; charset="UTF-8"',
  'Cookie': 'WMONID=sPDJXav32ei; lastAccess=1620358148921; globalDebug=false; JSESSIONID=QUdU1_F421fphwc0zH5JT6jayDaM5-GvN8yLfnjTPU75TZGY9K_9!-1435442688; JSESSIONID=m8BULWAFiVAIxwy-vcynVWzcuhNCXRPQk18DF4OycGxnfTA1ZRyD!-1270617111; WMONID=jzwO4AEb8k2',
  'Host': 'seibro.or.kr',
  'Origin': 'https://seibro.or.kr',
  'Referer': 'https://seibro.or.kr/websquare/control.jsp?w2xPath=/IPORTAL/user/derivCombi/BIP_CNTS07014V.xml&menuNo=901',
  'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
  'sec-ch-ua-mobile': '?0',
  'Sec-Fetch-Dest': 'empty',
  'Sec-Fetch-Mode': 'cors',
  'Sec-Fetch-Site': 'same-origin',
  'submissionid': 'submission_issucoTop10List',
  'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.85 Safari/537.36'
}

response = requests.request("POST", url, headers=headers, data=payload)

xmlString = response.text
jsonDump = json.dumps(xmltodict.parse(xmlString),ensure_ascii=False)
jsonText = json.loads(jsonDump)

wb = load_workbook(filename = '/Users/yerimstar/PycharmProjects/예탁원/템플릿/예탁원_발행동향(발행회사)_템플릿.xlsm', data_only=True)
ws = wb[wb.sheetnames[0]]
length = int(jsonText["vector"]["@result"])
if length > 1:
    for i in range(0, length):
        value = str(jsonText["vector"]["data"][i]["result"]["REP_SECN_NM"]["@value"])
        if i == length - 1:
            ws.cell(row=i + 3, column=1).value = ""
        else:
            ws.cell(row=i + 3, column=1).value = str(i + 1)
        ws.cell(row=i + 3, column=2).value = value
        ws.cell(row=i + 3, column=3).value = jsonText["vector"]["data"][i]["result"]["CNT_11"]["@value"]
        ws.cell(row=i + 3, column=4).value = jsonText["vector"]["data"][i]["result"]["CNT_21"]["@value"]
        ws.cell(row=i + 3, column=5).value = jsonText["vector"]["data"][i]["result"]["CNT_KRW"]["@value"]
        ws.cell(row=i + 3, column=6).value = jsonText["vector"]["data"][i]["result"]["CNT_OVERSEA"]["@value"]
        ws.cell(row=i + 3, column=7).value = jsonText["vector"]["data"][i]["result"]["CNT"]["@value"]
        val1 = jsonText["vector"]["data"][i]["result"]["SUMAMT_11"]["@value"]
        if len(val1) > 8:
            ws.cell(row=i + 3, column=8).value = float(val1[0:-8] +'.'+val1[-8])
        else:
            ws.cell(row=i + 3, column=8).value = int(val1)
        val2 = jsonText["vector"]["data"][i]["result"]["SUMAMT_21"]["@value"]
        if len(val2) > 8:
            ws.cell(row=i + 3, column=9).value = float(val2[0:-8] +'.'+val2[-8])
        else:
            ws.cell(row=i + 3, column=9).value = int(val2)
        val3 = jsonText["vector"]["data"][i]["result"]["SUMAMT_KRW"]["@value"]
        if len(val3) > 8:
            ws.cell(row=i + 3, column=10).value = float(val3[0:-8] +'.'+val3[-8])
        else:
            ws.cell(row=i + 3, column=10).value = int(val3)
        val4 = jsonText["vector"]["data"][i]["result"]["SUMAMT_OVERSEA"]["@value"]
        if len(val4) > 8:
            ws.cell(row=i + 3, column=11).value = float(val4[0:-8] +'.'+val4[-8])
        else:
            ws.cell(row=i + 3, column=11).value = int(val4)
        val5 = jsonText["vector"]["data"][i]["result"]["SUMAMT"]["@value"]
        if len(val5) > 8:
            ws.cell(row=i + 3, column=12).value = float(val5[0:-8] +'.'+val5[-8])
        else:
            ws.cell(row=i + 3, column=12).value = int(val5)
    wb.save("/Users/yerimstar/PycharmProjects/예탁원/결과물/예탁원_발행동향_Top10발행회사_" + value1 + "(" + date1 + "-" + date2 + ").xlsx")
else:
    print("값이 없습니다.")




