import requests
import json
import xmltodict
from openpyxl import load_workbook

date = input("기준 날짜를 입력하세요 ex) 20210501 : ")
url = "https://seibro.or.kr/websquare/engine/proworks/callServletService.jsp"

payload = "<reqParam action=\"stkslbsecnrankPList\" task=\"ksd.safe.bip.cnts.Loan.process.StkSlbSecnRankPTask\"><MENU_NO value=\"221\"/><CMM_BTN_ABBR_NM value=\"allview,allview,print,hwp,word,pdf,seach,xls,link,link,wide,wide,top,\"/><W2XPATH value=\"/IPORTAL/user/loan/BIP_CNTS08002V.xml\"/><sortStock value=\"0\"/><tDate value=\""+date+"\"/><START_PAGE value=\"1\"/><END_PAGE value=\"120\"/></reqParam>"
headers = {
  'Accept': 'application/xml',
  'Accept-Encoding': 'gzip, deflate, br',
  'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7,zh;q=0.6,zh-CN;q=0.5,zh-TW;q=0.4,zh-HK;q=0.3',
  'Connection': 'keep-alive',
  'Content-Type': 'application/xml; charset="UTF-8"',
  'Cookie': 'WMONID=sPDJXav32ei; lastAccess=1620358148921; globalDebug=false; JSESSIONID=pctZEm0s5sGIlrg-UwivLZTShXt4iw6NIMaGZaq1MA3vdZKJUk-m!-917870616; SeibroSLBPopup=done; JSESSIONID=bc5Y7wkkH3liLxGsIk2RQdt0MAgFs0Sgtv1iOdhGrGu-uxUHpt2Q!-917870616; WMONID=jzwO4AEb8k2',
  'Host': 'seibro.or.kr',
  'Origin': 'https://seibro.or.kr',
  'Referer': 'https://seibro.or.kr/websquare/control.jsp?w2xPath=/IPORTAL/user/loan/BIP_CNTS08002V.xml&menuNo=221',
  'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
  'sec-ch-ua-mobile': '?0',
  'Sec-Fetch-Dest': 'empty',
  'Sec-Fetch-Mode': 'cors',
  'Sec-Fetch-Site': 'same-origin',
  'submissionid': 'submission_stkslbsecnrankPList',
  'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.85 Safari/537.36'
}

response = requests.request("POST", url, headers=headers, data=payload)

xmlString = response.text
jsonDump = json.dumps(xmltodict.parse(xmlString),ensure_ascii=False)
jsonText = json.loads(jsonDump)

wb = load_workbook(filename = '/Users/yerimstar/PycharmProjects/예탁원/템플릿/예탁원_주식대차_대차종목순위_템플릿.xlsm', data_only=True)
ws = wb[wb.sheetnames[0]]
length = int(jsonText["vector"]["@result"])

if not length == 0:
    for i in range(0, length):
        ws.cell(row=i + 4, column=1).value = str(jsonText["vector"]["data"][i]["result"]["KOR_SECN_NM"]["@value"])
        ws.cell(row=i + 4, column=2).value = str(jsonText["vector"]["data"][i]["result"]["SHOTN_ISIN"]["@value"])
        ws.cell(row=i + 4, column=3).value = int(jsonText["vector"]["data"][i]["result"]["MATC_QTY"]["@value"])
        ws.cell(row=i + 4, column=4).value = int(jsonText["vector"]["data"][i]["result"]["RECALL_RED_QTY"]["@value"])
        ws.cell(row=i + 4, column=5).value = int(jsonText["vector"]["data"][i]["result"]["RED_QTY"]["@value"])
        ws.cell(row=i + 4, column=6).value = int(jsonText["vector"]["data"][i]["result"]["SLB_TR_REM_QTY"]["@value"])
        if int(jsonText["vector"]["data"][i]["result"]["REM_AMT"]["@value"][-6]) > 5:
            ws.cell(row=i + 4, column=7).value = int(
                jsonText["vector"]["data"][i]["result"]["REM_AMT"]["@value"][0:-6]) + 1
        else:
            ws.cell(row=i + 4, column=7).value = int(jsonText["vector"]["data"][i]["result"]["REM_AMT"]["@value"][0:-6])
    wb.save("/Users/yerimstar/PycharmProjects/예탁원/결과물/예탁원_대차종목별순위(" + date + ").xlsx")
else:
    print("조회되는 값이 없습니다.")